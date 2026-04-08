"""
EXCEL SPREADSHEET GENERATOR
Creates Ticket_Analysis.xlsx with ticket and feedbacks sheets,
including VLOOKUP, IF statements, and COUNTIFS formulas.
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    print("✓ openpyxl is available")
except ImportError:
    print("⚠ openpyxl not installed. Installing...")
    import subprocess
    import sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    print("✓ openpyxl installed successfully")


def create_ticket_analysis_excel():
    """Create the Excel workbook with ticket and feedback analysis."""
    
    # Create a new workbook
    wb = Workbook()
    
    # Remove default sheet and create two sheets
    wb.remove(wb.active)
    ticket_sheet = wb.create_sheet("ticket")
    feedback_sheet = wb.create_sheet("feedbacks")
    
    # ========================================================================
    # SHEET 1: TICKET DATA
    # ========================================================================
    
    # Headers
    ticket_sheet['A1'] = 'ticket_id'
    ticket_sheet['B1'] = 'created_at'
    ticket_sheet['C1'] = 'updated_at'
    ticket_sheet['D1'] = 'status'
    
    # Sample data
    ticket_data = [
        (1, '2021-11-10 09:30:00', '2021-11-10 10:15:00', 'Closed'),
        (2, '2021-11-10 14:00:00', '2021-11-10 15:45:00', 'Closed'),
        (3, '2021-11-11 08:00:00', '2021-11-12 09:00:00', 'Closed'),
        (4, '2021-11-15 10:30:00', '2021-11-15 11:00:00', 'Closed'),
        (5, '2021-11-15 13:45:00', '2021-11-16 08:00:00', 'Closed'),
    ]
    
    for idx, (ticket_id, created, updated, status) in enumerate(ticket_data, start=2):
        ticket_sheet[f'A{idx}'] = ticket_id
        ticket_sheet[f'B{idx}'] = created
        ticket_sheet[f'C{idx}'] = updated
        ticket_sheet[f'D{idx}'] = status
    
    # Format ticket sheet
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ['A1', 'B1', 'C1', 'D1']:
        ticket_sheet[cell].fill = header_fill
        ticket_sheet[cell].font = header_font
        ticket_sheet[cell].alignment = Alignment(horizontal="center", vertical="center")
    
    # Set column widths
    ticket_sheet.column_dimensions['A'].width = 12
    ticket_sheet.column_dimensions['B'].width = 20
    ticket_sheet.column_dimensions['C'].width = 20
    ticket_sheet.column_dimensions['D'].width = 12
    
    # ========================================================================
    # SHEET 2: FEEDBACK DATA WITH FORMULAS
    # ========================================================================
    
    # Headers
    feedback_sheet['A1'] = 'feedback_id'
    feedback_sheet['B1'] = 'ticket_id'
    feedback_sheet['C1'] = 'created_at (VLOOKUP)'
    feedback_sheet['D1'] = 'resolved_at'
    feedback_sheet['E1'] = 'Same Day?'
    feedback_sheet['F1'] = 'Same Hour?'
    
    # Apply header formatting
    for cell in ['A1', 'B1', 'C1', 'D1', 'E1', 'F1']:
        feedback_sheet[cell].fill = header_fill
        feedback_sheet[cell].font = header_font
        feedback_sheet[cell].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Sample feedback data
    feedback_data = [
        (1, 1, '2021-11-10 10:30:00'),
        (2, 2, '2021-11-10 16:00:00'),
        (3, 3, '2021-11-12 09:30:00'),
        (4, 4, '2021-11-15 11:15:00'),
        (5, 5, '2021-11-16 09:00:00'),
    ]
    
    for idx, (feedback_id, ticket_id, resolved_at) in enumerate(feedback_data, start=2):
        feedback_sheet[f'A{idx}'] = feedback_id
        feedback_sheet[f'B{idx}'] = ticket_id
        # Formula: VLOOKUP to get created_at from ticket sheet
        feedback_sheet[f'C{idx}'] = f'=VLOOKUP(B{idx},ticket!A:B,2,FALSE)'
        feedback_sheet[f'D{idx}'] = resolved_at
        # Formula: Check if same day
        feedback_sheet[f'E{idx}'] = f'=IF(INT(C{idx})=INT(D{idx}),"Yes","No")'
        # Formula: Check if same hour
        feedback_sheet[f'F{idx}'] = f'=IF(HOUR(C{idx})=HOUR(D{idx}),"Yes","No")'
    
    # Add summary row with COUNTIFS
    summary_row = len(feedback_data) + 3
    feedback_sheet[f'A{summary_row}'] = 'Summary:'
    feedback_sheet[f'A{summary_row}'].font = Font(bold=True)
    
    feedback_sheet[f'C{summary_row}'] = 'Resolved Same Day & Hour:'
    feedback_sheet[f'C{summary_row}'].font = Font(bold=True)
    
    # COUNTIFS formula: Count rows where E="Yes" AND F="Yes"
    data_end = len(feedback_data) + 1
    feedback_sheet[f'D{summary_row}'] = f'=COUNTIFS(E2:E{data_end},"Yes",F2:F{data_end},"Yes")'
    feedback_sheet[f'D{summary_row}'].font = Font(bold=True, color="FF0000")
    
    # Set column widths for feedback sheet
    feedback_sheet.column_dimensions['A'].width = 12
    feedback_sheet.column_dimensions['B'].width = 12
    feedback_sheet.column_dimensions['C'].width = 22
    feedback_sheet.column_dimensions['D'].width = 22
    feedback_sheet.column_dimensions['E'].width = 12
    feedback_sheet.column_dimensions['F'].width = 12
    
    # Format data cells
    for row in range(2, data_end + 1):
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            feedback_sheet[f'{col}{row}'].alignment = Alignment(horizontal="center")
    
    # ========================================================================
    # Save workbook
    # ========================================================================
    
    output_file = r"h:\hotel-menagement\PlatinumRx_Assignment\Spreadsheets\Ticket_Analysis.xlsx"
    wb.save(output_file)
    
    return output_file


if __name__ == "__main__":
    print("\n" + "=" * 70)
    print("CREATING EXCEL SPREADSHEET")
    print("=" * 70 + "\n")
    
    try:
        output_file = create_ticket_analysis_excel()
        print(f"✓ Excel file created successfully!")
        print(f"  Location: {output_file}")
        print(f"\n  Sheets created:")
        print(f"    • ticket: Contains ticket_id, created_at, updated_at, status")
        print(f"    • feedbacks: Contains feedback data with formulas")
        print(f"\n  Formulas included:")
        print(f"    • VLOOKUP: Fetches created_at from ticket sheet")
        print(f"    • IF: Checks if same day (compares dates)")
        print(f"    • IF: Checks if same hour (compares hours)")
        print(f"    • COUNTIFS: Counts resolved same-day AND same-hour tickets")
        print(f"\n" + "=" * 70)
    
    except Exception as e:
        print(f"✗ Error creating Excel file:")
        print(f"  {type(e).__name__}: {e}")
        print(f"\n  Make sure openpyxl is installed:")
        print(f"  pip install openpyxl")
